from pathlib import Path
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import Http404
from django.shortcuts import get_object_or_404, redirect, render

from openpyxl import load_workbook
from .forms import EditDocumentForm, RegisterForm, UploadForm
from .models import Document


UPLOAD_SUCCESS = "✓ Upload thành công"
DELETE_SUCCESS = "✓ Xóa thành công"
UPDATE_SUCCESS = "✓ Cập nhật thành công"
APPROVE_SUCCESS = "✓ Duyệt thành công"

DOCX_NAMESPACE = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def get_file_extension(document):
    return Path(document.file.name).suffix.lower()


def extract_docx_text(document):
    try:
        with ZipFile(document.file.path) as docx:
            xml_content = docx.read("word/document.xml")
    except (BadZipFile, KeyError, FileNotFoundError):
        return ""

    root = ElementTree.fromstring(xml_content)
    paragraphs = []

    for paragraph in root.iter(f"{DOCX_NAMESPACE}p"):
        text_parts = [
            text.text
            for text in paragraph.iter(f"{DOCX_NAMESPACE}t")
            if text.text
        ]

        if text_parts:
            paragraphs.append("".join(text_parts))

    return "\n".join(paragraphs)

def extract_xlsx_data(document):
    try:
        workbook = load_workbook(document.file.path, data_only=True)
        sheet = workbook.active

        rows = []

        for row in sheet.iter_rows(max_row=50, values_only=True):
            rows.append(row)

        return rows

    except Exception:
        return []

def get_document_preview(document):
    file_extension = get_file_extension(document)

    if file_extension == ".docx":
        return {
            "type": "docx",
            "text": extract_docx_text(document),
        }

    if file_extension == ".xlsx":
        return {
            "type": "xlsx",
            "rows": extract_xlsx_data(document),
        }

    if file_extension == ".pdf":
        return {"type": "pdf"}

    if file_extension in [".png", ".jpg", ".jpeg", ".gif", ".webp"]:
        return {"type": "image"}

    return {"type": "unsupported"}
def home(request):
    docs = Document.objects.filter(approved=True).order_by("-uploaded_at")
    q = request.GET.get("q")

    if q:
        docs = docs.filter(
            Q(title__icontains=q) | Q(uploaded_by__username__icontains=q)
        )

    paginator = Paginator(docs, 10)
    docs = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "home.html",
        {
            "docs": docs,
            "q": q,
        },
    )


@login_required
def upload_document(request):
    if request.method == "POST":
        form = UploadForm(request.POST, request.FILES)

        if form.is_valid():
            Document.objects.create(
                title=form.cleaned_data["title"],
                file=request.FILES["file"],
                uploaded_by=request.user,
            )
            messages.success(request, UPLOAD_SUCCESS)

            return redirect("/")
    else:
        form = UploadForm()

    return render(request, "upload.html", {"form": form})


def register(request):
    if request.method == "POST":
        form = RegisterForm(request.POST)

        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data["password"])
            user.save()
            login(request, user)

            return redirect("/")
    else:
        form = RegisterForm()

    return render(request, "register.html", {"form": form})


@login_required
def my_documents(request):
    docs = Document.objects.filter(uploaded_by=request.user)

    return render(request, "my_documents.html", {"docs": docs})


def document_detail(request, id):
    doc = get_object_or_404(Document, id=id)

    can_view_pending = (
        request.user.is_authenticated
        and (request.user.is_superuser or doc.uploaded_by == request.user)
    )

    if not doc.approved and not can_view_pending:
        raise Http404

    return render(
        request,
        "document_detail.html",
        {
            "doc": doc,
            "preview": get_document_preview(doc),
        },
    )


@login_required
def delete_document(request, id):
    doc = get_object_or_404(
        Document,
        id=id,
        uploaded_by=request.user,
    )

    if request.method == "POST":
        doc.file.delete()
        doc.delete()
        messages.success(request, DELETE_SUCCESS)

        return redirect("/my-documents/")

    return render(request, "confirm_delete.html", {"doc": doc})


@login_required
def edit_document(request, id):
    doc = get_object_or_404(
        Document,
        id=id,
        uploaded_by=request.user,
    )

    if request.method == "POST":
        form = EditDocumentForm(request.POST, instance=doc)

        if form.is_valid():
            document = form.save(commit=False)
            document.approved = False
            document.save()
            messages.success(request, UPDATE_SUCCESS)

            return redirect("/my-documents/")
    else:
        form = EditDocumentForm(instance=doc)

    return render(request, "edit_document.html", {"form": form})


@login_required
def dashboard(request):
    if not request.user.is_superuser:
        return redirect("/")

    pending_docs = Document.objects.filter(approved=False).order_by("-uploaded_at")
    approved_docs = Document.objects.filter(approved=True).order_by("-uploaded_at")

    return render(
        request,
        "dashboard.html",
        {
            "total_documents": Document.objects.count(),
            "pending_documents": pending_docs.count(),
            "approved_documents": approved_docs.count(),
            "total_users": User.objects.count(),
            "pending_docs": pending_docs,
            "approved_docs": approved_docs,
        },
    )


@login_required
def approve_document(request, id):
    if not request.user.is_superuser:
        return redirect("/")

    doc = get_object_or_404(Document, id=id)
    doc.approved = True
    doc.save()
    messages.success(request, APPROVE_SUCCESS)

    return redirect("/dashboard/")


@login_required
def admin_delete_document(request, id):
    if not request.user.is_superuser:
        return redirect("/")

    doc = get_object_or_404(Document, id=id)
    doc.file.delete()
    doc.delete()
    messages.success(request, DELETE_SUCCESS)

    return redirect("/dashboard/")
